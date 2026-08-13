"""
从品牌图片 Excel 下载 YOLOE visual prompt 参考图。

该脚本服务于 YOLOE visual prompt 预标注流程：
- 从 Excel 中读取品牌列和图片 URL 列。
- 按品牌把参考包装图下载到 visual_prompts/<品牌>/。
- 对同一个 URL 做全局去重，避免重复下载同一张参考图。
- 下载后验证图片完整性，并生成 JSON/CSV 下载报告。

默认输入是用户提供的品牌图片 Excel：
/Users/guobiao/Downloads/品牌图片_1786521889720.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import mimetypes
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from PIL import Image

# 项目根目录：scripts/data_import/<当前文件> 向上两级即项目根目录。
PROJECT_ROOT = Path(__file__).resolve().parents[2]
# 用户提供的品牌参考图 Excel 默认路径。
DEFAULT_EXCEL = Path("/Users/guobiao/Downloads/品牌图片_1786521889720.xlsx")
# Excel 中品牌列与图片 URL 列的默认列名。
DEFAULT_BRAND_COLUMN = "brand"
DEFAULT_URL_COLUMN = "attach_file"
# YOLOE visual prompt 默认参考图根目录。
DEFAULT_OUTPUT_ROOT = PROJECT_ROOT / "datasets" / "multibrand" / "visual_prompts"
# 下载报告默认目录，放在参考图根目录下的 metadata 中。
DEFAULT_METADATA_DIR = DEFAULT_OUTPUT_ROOT / "metadata"
# 本地允许落盘和识别的图片后缀。
IMAGE_EXTENSIONS = {".bmp", ".jpeg", ".jpg", ".png", ".tif", ".tiff", ".webp"}
# 常见图片 Content-Type 到扩展名的映射，用于 URL 没有后缀的情况。
CONTENT_TYPE_EXTENSIONS = {
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
    "image/bmp": ".bmp",
    "image/tiff": ".tiff",
}
# URL 下载时使用轻量 User-Agent，降低部分对象存储拒绝默认 Python UA 的概率。
REQUEST_HEADERS = {"User-Agent": "yolo-example-visual-prompt-import/1.0"}


@dataclass(frozen=True)
class VisualPromptImageRecord:
    """一张待下载品牌参考图在 Excel 中的来源记录。"""

    row_number: int  # Excel 行号，用于回溯原始数据。
    url_index: int  # 同一 Excel 单元格中第几个 URL，用于避免同一行多图重名。
    brand: str  # 已清洗后的品牌目录名。
    url: str  # 图片 HTTP(S) 地址。


@dataclass(frozen=True)
class VisualPromptDownloadResult:
    """单张品牌参考图的下载结果，用于写入 JSON/CSV 报告。"""

    row_number: int  # Excel 行号。
    url_index: int  # 同一单元格中的 URL 序号。
    brand: str  # 品牌目录名。
    url: str  # 图片 HTTP(S) 地址。
    status: str  # downloaded、skipped 或 failed。
    path: str = ""  # downloaded/skipped 时的本地文件路径。
    error: str = ""  # failed 时的错误原因。


def find_column_index(headers: list[object], column_name: str) -> int:
    """
    在 Excel 表头中查找指定列名的位置。

    表头来自 openpyxl，可能包含 None 或数字；这里统一转成去空白字符串，
    保证列名匹配行为和普通业务 Excel 的肉眼显示一致。
    """
    normalized = [str(value).strip() if value is not None else "" for value in headers]
    if column_name not in normalized:
        available = ", ".join(header for header in normalized if header)
        raise ValueError(f"Excel 中找不到列：{column_name}。当前列：{available}")
    return normalized.index(column_name)


def split_urls(value: object) -> list[str]:
    """
    从一个单元格中拆分图片 URL。

    业务 Excel 中一格可能放多张图片，常见分隔符包括换行、英文逗号、中文逗号
    和分号。拆分后会去掉空片段，但不在这里判断 URL 合法性。
    """
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    for separator in ("\n", "\r", ";", "，", ","):
        text = text.replace(separator, " ")
    return [part.strip() for part in text.split() if part.strip()]


def sanitize_brand_dir_name(value: object) -> str:
    """
    把 Excel 品牌值清洗成安全的目录名。

    允许保留字母、数字、点、下划线和连字符；其它字符统一替换为下划线。
    这样既能兼容 "T-GUARD" 这类品牌名，也能避免斜杠等字符意外创建多级目录。
    """
    if value is None:
        raise ValueError("品牌为空")
    raw_name = str(value).strip()
    if not raw_name:
        raise ValueError("品牌为空")
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", raw_name).strip("._-")
    if not safe_name:
        raise ValueError(f"品牌目录名无效：{value}")
    return safe_name


def is_http_url(url: str) -> bool:
    """判断字符串是否是可下载的 HTTP(S) URL。"""
    parsed = urlparse(url)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def read_visual_prompt_records(
    excel_path: Path,
    brand_column: str = DEFAULT_BRAND_COLUMN,
    url_column: str = DEFAULT_URL_COLUMN,
    limit: int | None = None,
) -> list[VisualPromptImageRecord]:
    """
    从 Excel 中读取品牌参考图下载记录。

    读取阶段会做三件轻量清洗：
    1. 空品牌跳过，因为无法决定目标品牌目录。
    2. 非 HTTP(S) URL 跳过，避免把备注文字当成图片地址。
    3. 对完整 URL 做全局去重，避免同一张参考图重复下载。
    """
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration as exc:
            raise ValueError("Excel 是空文件。") from exc

        return records_from_rows(
            rows=rows,
            brand_index=find_column_index(headers, brand_column),
            url_index=find_column_index(headers, url_column),
            limit=limit,
        )
    finally:
        workbook.close()


def records_from_rows(
    rows: object,
    brand_index: int,
    url_index: int,
    limit: int | None,
) -> list[VisualPromptImageRecord]:
    """
    从 openpyxl 行迭代器中提取已清洗、已去重的品牌参考图记录。

    单独拆出该函数可以让 Excel 打开/关闭逻辑和逐行清洗逻辑保持简单，
    也便于后续如需增加“跳过原因报告”时只改这里。
    """
    records: list[VisualPromptImageRecord] = []
    seen_urls: set[str] = set()
    for row_number, row in enumerate(rows, start=2):
        raw_brand = row[brand_index] if brand_index < len(row) else None
        try:
            brand = sanitize_brand_dir_name(raw_brand)
        except ValueError:
            # 空品牌行无法归类到 visual_prompts/<brand>/，直接跳过。
            continue

        raw_urls = row[url_index] if url_index < len(row) else None
        for cell_url_index, url in enumerate(split_urls(raw_urls), start=1):
            if not is_http_url(url) or url in seen_urls:
                continue
            seen_urls.add(url)
            records.append(
                VisualPromptImageRecord(
                    row_number=row_number,
                    url_index=cell_url_index,
                    brand=brand,
                    url=url,
                )
            )
            if limit is not None and len(records) >= limit:
                return records
    return records


def extension_from_response(url: str, content_type: str | None) -> str:
    """
    根据 URL 路径或 HTTP Content-Type 推断图片后缀。

    优先使用 URL 自带后缀，因为对象存储链接通常保留原始文件名；如果 URL 没有
    图片后缀，再尝试读取响应头；最后回退为 .jpg。
    """
    parsed_suffix = Path(urlparse(url).path).suffix.lower()
    if parsed_suffix in IMAGE_EXTENSIONS:
        return parsed_suffix
    if content_type:
        media_type = content_type.split(";", 1)[0].strip().lower()
        if media_type in CONTENT_TYPE_EXTENSIONS:
            return CONTENT_TYPE_EXTENSIONS[media_type]
        guessed = mimetypes.guess_extension(media_type)
        if guessed in IMAGE_EXTENSIONS:
            return guessed
    return ".jpg"


def attachment_stem_from_url(url: str) -> str:
    """
    从 URL 路径中提取原附件名称 stem，并清洗成本地文件名片段。

    如果远端路径为空或文件名全是非法字符，则使用 attachment 兜底。
    """
    parsed_name = Path(urlparse(url).path).name
    stem = Path(parsed_name).stem.strip()
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._-")
    return safe_stem or "attachment"


def normalize_extension(extension: str) -> str:
    """确保扩展名以点开头，并统一转为小写。"""
    clean_extension = extension.lower().strip()
    if not clean_extension.startswith("."):
        clean_extension = f".{clean_extension}"
    return clean_extension


def make_filename(record: VisualPromptImageRecord, extension: str) -> str:
    """
    生成品牌参考图文件名。

    格式：row<Excel行号>_<URL序号>_<原附件名>.<后缀>。
    URL 序号可以避免同一行多个附件恰好同名时互相覆盖。
    """
    safe_extension = normalize_extension(extension)
    attachment_stem = attachment_stem_from_url(record.url)
    return (
        f"row{record.row_number:05d}_{record.url_index:02d}_"
        f"{attachment_stem}{safe_extension}"
    )


def target_path_for_record(
    record: VisualPromptImageRecord,
    output_root: Path,
    extension: str,
) -> Path:
    """返回一条记录在 visual_prompts/<品牌>/ 下的目标保存路径。"""
    brand_dir = output_root / sanitize_brand_dir_name(record.brand)
    return brand_dir / make_filename(record, extension)


def existing_download_path(record: VisualPromptImageRecord, output_root: Path) -> Path | None:
    """
    查找某条记录是否已经下载过。

    扩展名可能由远端 Content-Type 决定，因此这里只匹配固定文件名前缀和允许的图片后缀。
    如果匹配到多个文件，说明目录中存在歧义，返回 None 让下载流程重新决策。
    """
    brand_dir = output_root / sanitize_brand_dir_name(record.brand)
    filename_prefix = (
        f"row{record.row_number:05d}_{record.url_index:02d}_"
        f"{attachment_stem_from_url(record.url)}"
    )
    matches = [
        path
        for path in brand_dir.glob(f"{filename_prefix}.*")
        if path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS
    ]
    return matches[0] if len(matches) == 1 else None


def completed_download_results(
    records: list[VisualPromptImageRecord],
    output_root: Path,
) -> list[VisualPromptDownloadResult] | None:
    """
    如果所有 Excel 记录都已有本地图片，返回 skipped 结果；否则返回 None。

    该函数用于主流程的快速恢复：全量已完成时无需创建线程，也无需访问网络。
    """
    results: list[VisualPromptDownloadResult] = []
    for record in records:
        existing_path = existing_download_path(record, output_root)
        if existing_path is None:
            return None
        results.append(
            VisualPromptDownloadResult(
                row_number=record.row_number,
                url_index=record.url_index,
                brand=record.brand,
                url=record.url,
                status="skipped",
                path=str(existing_path),
            )
        )
    return results


def validate_image_file(path: Path) -> None:
    """
    验证下载后的文件是否能被 PIL 识别为完整图片。

    PIL 的 verify() 不会解码全部像素，但能快速发现截断、HTML 错误页等明显异常。
    """
    with Image.open(path) as image:
        image.verify()


def result_for_record(
    record: VisualPromptImageRecord,
    status: str,
    path: str = "",
    error: str = "",
) -> VisualPromptDownloadResult:
    """用一条 Excel 记录生成下载结果，避免多处重复传递相同字段。"""
    return VisualPromptDownloadResult(
        row_number=record.row_number,
        url_index=record.url_index,
        brand=record.brand,
        url=record.url,
        status=status,
        path=path,
        error=error,
    )


def download_missing_record(
    record: VisualPromptImageRecord,
    output_root: Path,
    timeout: int,
) -> VisualPromptDownloadResult:
    """下载尚未存在于本地的参考图记录。"""
    request = Request(record.url, headers=REQUEST_HEADERS)
    with urlopen(request, timeout=timeout) as response:  # noqa: S310 - 用户 Excel 图片 URL。
        content_type = response.headers.get("Content-Type")
        extension = extension_from_response(record.url, content_type)
        target = target_path_for_record(record, output_root, extension)
        target.parent.mkdir(parents=True, exist_ok=True)
        if target.exists():
            return result_for_record(record, "skipped", path=str(target))
        content = response.read()

    if not content:
        return result_for_record(record, "failed", error="empty response")

    target.write_bytes(content)
    try:
        validate_image_file(target)
    except Exception as exc:  # pylint: disable=broad-exception-caught
        target.unlink(missing_ok=True)
        return result_for_record(record, "failed", error=f"invalid image: {exc}")
    return result_for_record(record, "downloaded", path=str(target))


def download_one(
    record: VisualPromptImageRecord,
    output_root: Path,
    timeout: int,
) -> VisualPromptDownloadResult:
    """下载单张品牌参考图，并返回结构化结果。"""
    try:
        existing_path = existing_download_path(record, output_root)
        if existing_path is not None:
            result = result_for_record(record, "skipped", path=str(existing_path))
        else:
            result = download_missing_record(record, output_root, timeout)
    except HTTPError as exc:
        result = result_for_record(record, "failed", error=f"HTTP {exc.code}")
    except URLError as exc:
        result = result_for_record(record, "failed", error=str(exc.reason))
    except Exception as exc:  # pylint: disable=broad-exception-caught
        result = result_for_record(record, "failed", error=str(exc))
    return result


def write_reports(results: list[VisualPromptDownloadResult], metadata_dir: Path) -> None:
    """把下载结果写成 JSON 和 CSV，方便人工查看与后续审计。"""
    metadata_dir.mkdir(parents=True, exist_ok=True)
    rows = [asdict(result) for result in results]
    json_path = metadata_dir / "download_report.json"
    csv_path = metadata_dir / "download_report.csv"

    json_path.write_text(
        json.dumps(rows, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    with csv_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "row_number",
                "url_index",
                "brand",
                "url",
                "status",
                "path",
                "error",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)


def positive_int(value: str) -> int:
    """argparse 类型转换：只接受大于 0 的整数。"""
    number = int(value)
    if number < 1:
        raise argparse.ArgumentTypeError("必须大于 0")
    return number


def parse_args() -> argparse.Namespace:
    """解析命令行参数。"""
    parser = argparse.ArgumentParser(description="从品牌图片 Excel 下载 YOLOE visual prompt 参考图。")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="品牌图片 Excel 文件路径")
    parser.add_argument("--brand-column", default=DEFAULT_BRAND_COLUMN, help="品牌列名")
    parser.add_argument("--url-column", default=DEFAULT_URL_COLUMN, help="图片 URL 列名")
    parser.add_argument(
        "--output-root",
        type=Path,
        default=DEFAULT_OUTPUT_ROOT,
        help="visual_prompts 根目录",
    )
    parser.add_argument("--metadata-dir", type=Path, default=DEFAULT_METADATA_DIR, help="下载报告目录")
    parser.add_argument("--workers", type=positive_int, default=8, help="并发下载线程数")
    parser.add_argument(
        "--timeout",
        type=positive_int,
        default=30,
        help="单张图片下载超时时间（秒）",
    )
    parser.add_argument(
        "--limit",
        type=positive_int,
        default=None,
        help="最多读取多少个去重后的 URL，调试用",
    )
    return parser.parse_args()


def main() -> None:
    """脚本主入口：读取 Excel、下载参考图、写报告并打印汇总。"""
    args = parse_args()
    if not args.excel.is_file():
        raise SystemExit(f"Excel 文件不存在：{args.excel}")

    args.output_root.mkdir(parents=True, exist_ok=True)
    records = read_visual_prompt_records(
        excel_path=args.excel,
        brand_column=args.brand_column,
        url_column=args.url_column,
        limit=args.limit,
    )
    if not records:
        raise SystemExit(
            f"Excel 列 {args.brand_column}/{args.url_column} 中没有可下载的品牌 HTTP(S) 图片 URL。"
        )

    results = completed_download_results(records, args.output_root)
    if results is not None:
        write_reports(results, args.metadata_dir)
        print(f"全部已存在，跳过下载：total={len(results)}")
        print(f"参考图目录：{args.output_root}")
        print(f"下载报告：{args.metadata_dir / 'download_report.csv'}")
        return

    results = []
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = [
            executor.submit(download_one, record, args.output_root, args.timeout)
            for record in records
        ]
        for index, future in enumerate(as_completed(futures), start=1):
            result = future.result()
            results.append(result)
            location = result.path or result.error
            print(
                f"[{index}/{len(futures)}] {result.status}: "
                f"row {result.row_number} url#{result.url_index} {result.brand} {location}"
            )

    results.sort(key=lambda item: (item.brand, item.row_number, item.url_index, item.url))
    write_reports(results, args.metadata_dir)

    downloaded = sum(result.status == "downloaded" for result in results)
    skipped = sum(result.status == "skipped" for result in results)
    failed = sum(result.status == "failed" for result in results)
    print(f"完成：downloaded={downloaded}, skipped={skipped}, failed={failed}")
    print(f"参考图目录：{args.output_root}")
    print(f"下载报告：{args.metadata_dir / 'download_report.csv'}")


if __name__ == "__main__":
    main()
