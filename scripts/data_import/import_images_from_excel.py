"""
Excel 图片下载脚本。

该脚本从 Excel 文件中读取图片 URL 并批量下载，用于数据收集流程：
- 支持从 Excel 指定列读取图片 URL
- 支持多线程并发下载
- 自动检测图片格式并生成合适的文件名
- 验证下载的图片完整性
- 生成详细的下载报告（JSON 和 CSV）
"""

from __future__ import annotations

import argparse
import csv
import json
import mimetypes
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from PIL import Image

# 项目根目录
PROJECT_ROOT = Path(__file__).resolve().parents[2]
# 默认 Excel 文件路径
DEFAULT_EXCEL = Path("/Users/guobiao/Downloads/8e96894159cc584f0c7a27faaa4acc45.xlsx")
# 默认图片输出目录
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "datasets" / "multibrand" / "raw" / "images"
# 默认元数据输出目录
DEFAULT_METADATA_DIR = PROJECT_ROOT / "datasets" / "multibrand" / "raw" / "metadata"
# 默认读取的列名
DEFAULT_COLUMN = "整改后图片URL"
# 支持的图片格式后缀
IMAGE_EXTENSIONS = {".bmp", ".jpeg", ".jpg", ".png", ".tif", ".tiff", ".webp"}
# MIME 类型到文件后缀的映射
CONTENT_TYPE_EXTENSIONS = {
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
    "image/bmp": ".bmp",
    "image/tiff": ".tiff",
}


@dataclass(frozen=True)
class ImageRecord:
    """图片记录数据类，包含行号和 URL。"""
    row_number: int  # Excel 中的行号（从 2 开始，因为第 1 行是表头）
    url: str  # 图片 URL


@dataclass(frozen=True)
class DownloadResult:
    """下载结果数据类。"""
    row_number: int  # Excel 中的行号
    url: str  # 图片 URL
    status: str  # 下载状态："downloaded"、"skipped" 或 "failed"
    path: str = ""  # 下载成功时的文件路径
    error: str = ""  # 下载失败时的错误信息


def find_column_index(headers: list[object], column_name: str) -> int:
    """
    在表头中查找指定列的索引。
    
    Args:
        headers: 表头列表
        column_name: 要查找的列名
    
    Returns:
        列索引（从 0 开始）
    
    Raises:
        ValueError: 找不到指定列
    """
    # 标准化表头（去除空格，转换为字符串）
    normalized = [str(value).strip() if value is not None else "" for value in headers]
    if column_name not in normalized:
        available = ", ".join(header for header in normalized if header)
        raise ValueError(f"Excel 中找不到列：{column_name}。当前列：{available}")
    return normalized.index(column_name)


def split_urls(value: object) -> list[str]:
    """
    从单元格值中提取 URL 列表。
    
    支持多种分隔符：换行符、分号、逗号等。
    
    Args:
        value: 单元格值
    
    Returns:
        URL 列表
    """
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    # 将各种分隔符统一替换为空格
    for separator in ("\n", "\r", ";", "，", ","):
        text = text.replace(separator, " ")
    # 分割并过滤空字符串
    return [part.strip() for part in text.split() if part.strip()]


def read_image_urls(excel_path: Path, column_name: str) -> list[ImageRecord]:
    """
    从 Excel 文件中读取图片 URL。
    
    Args:
        excel_path: Excel 文件路径
        column_name: 包含图片 URL 的列名
    
    Returns:
        图片记录列表（已去重）
    
    Raises:
        ValueError: Excel 为空或找不到指定列
    """
    # 以只读模式打开 Excel
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = list(next(rows))
    except StopIteration as exc:
        raise ValueError("Excel 是空文件。") from exc

    # 查找目标列的索引
    column_index = find_column_index(headers, column_name)
    records: list[ImageRecord] = []
    seen_urls: set[str] = set()  # 用于去重
    # 遍历数据行（从第 2 行开始）
    for row_number, row in enumerate(rows, start=2):
        value = row[column_index] if column_index < len(row) else None
        # 提取 URL
        for url in split_urls(value):
            parsed = urlparse(url)
            # 只保留有效的 HTTP(S) URL
            if parsed.scheme not in {"http", "https"} or not parsed.netloc:
                continue
            # 去重
            if url in seen_urls:
                continue
            seen_urls.add(url)
            records.append(ImageRecord(row_number=row_number, url=url))
    return records


def extension_from_response(url: str, content_type: str | None) -> str:
    """
    从 URL 或 Content-Type 推断图片文件后缀。
    
    优先级：
    1. URL 路径中的后缀
    2. Content-Type 头部
    3. 默认使用 .jpg
    
    Args:
        url: 图片 URL
        content_type: HTTP 响应的 Content-Type 头部
    
    Returns:
        文件后缀（如 ".jpg"）
    """
    # 尝试从 URL 路径中提取后缀
    parsed_suffix = Path(urlparse(url).path).suffix.lower()
    if parsed_suffix in IMAGE_EXTENSIONS:
        return parsed_suffix
    # 尝试从 Content-Type 推断
    if content_type:
        media_type = content_type.split(";", 1)[0].strip().lower()
        if media_type in CONTENT_TYPE_EXTENSIONS:
            return CONTENT_TYPE_EXTENSIONS[media_type]
        guessed = mimetypes.guess_extension(media_type)
        if guessed in IMAGE_EXTENSIONS:
            return guessed
    # 默认使用 .jpg
    return ".jpg"


def attachment_stem_from_url(url: str) -> str:
    """从 URL 路径中提取原附件文件名 stem。"""
    parsed_name = Path(urlparse(url).path).name
    stem = Path(parsed_name).stem.strip()
    # 只保留常见安全字符，避免 URL 中异常字符影响本地文件名。
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._-")
    return safe_stem or "attachment"


def make_filename(record: ImageRecord, extension: str) -> str:
    """
    生成文件名：行号 + 原附件名称。

    原附件名称本身就是业务侧生成的唯一 hash，因此不再对完整 URL 重新计算 hash。

    Args:
        record: 图片记录
        extension: 文件后缀

    Returns:
        文件名（如 "row00002_67cfc8156160c2fd227aef004b771854.webp"）
    """
    attachment_stem = attachment_stem_from_url(record.url)
    return f"row{record.row_number:05d}_{attachment_stem}{extension}"


def download_one(record: ImageRecord, output_dir: Path, timeout: int) -> DownloadResult:
    """
    下载单张图片。
    
    Args:
        record: 图片记录
        output_dir: 输出目录
        timeout: 下载超时时间（秒）
    
    Returns:
        下载结果
    """
    try:
        # 构建 HTTP 请求
        request = Request(record.url, headers={"User-Agent": "softcare-yolo-demo/0.1"})
        with urlopen(request, timeout=timeout) as response:
            if response.status != 200:
                return DownloadResult(record.row_number, record.url, "failed", error=f"HTTP {response.status}")
            content_type = response.headers.get("Content-Type")
            # 推断文件后缀
            extension = extension_from_response(record.url, content_type)
            target = output_dir / make_filename(record, extension)
            # 如果文件已存在，跳过
            if target.exists():
                return DownloadResult(record.row_number, record.url, "skipped", path=str(target))
            content = response.read()

        # 检查响应内容
        if not content:
            return DownloadResult(record.row_number, record.url, "failed", error="empty response")

        # 写入文件
        target.write_bytes(content)
        # 验证图片完整性
        try:
            with Image.open(target) as image:
                image.verify()
        except Exception as exc:
            target.unlink(missing_ok=True)
            return DownloadResult(record.row_number, record.url, "failed", error=f"invalid image: {exc}")
        return DownloadResult(record.row_number, record.url, "downloaded", path=str(target))
    except HTTPError as exc:
        return DownloadResult(record.row_number, record.url, "failed", error=f"HTTP {exc.code}")
    except URLError as exc:
        return DownloadResult(record.row_number, record.url, "failed", error=str(exc.reason))
    except Exception as exc:
        return DownloadResult(record.row_number, record.url, "failed", error=str(exc))


def write_reports(results: list[DownloadResult], metadata_dir: Path) -> None:
    """
    生成下载报告（JSON 和 CSV）。
    
    Args:
        results: 下载结果列表
        metadata_dir: 元数据输出目录
    """
    metadata_dir.mkdir(parents=True, exist_ok=True)
    json_path = metadata_dir / "download_report.json"
    csv_path = metadata_dir / "download_report.csv"

    rows = [result.__dict__ for result in results]
    # 保存 JSON 报告
    json_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    # 保存 CSV 报告
    with csv_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=["row_number", "url", "status", "path", "error"])
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    """Excel 图片下载脚本主入口。"""
    parser = argparse.ArgumentParser(description="从 Excel 的图片 URL 列下载未标注原图。")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="Excel 文件路径")
    parser.add_argument("--column", default=DEFAULT_COLUMN, help="图片 URL 列名")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="原图保存目录")
    parser.add_argument("--metadata-dir", type=Path, default=DEFAULT_METADATA_DIR, help="下载报告目录")
    parser.add_argument("--workers", type=int, default=8, help="并发下载线程数")
    parser.add_argument("--timeout", type=int, default=30, help="单张图片下载超时时间（秒）")
    args = parser.parse_args()

    # 验证参数
    if not args.excel.is_file():
        raise SystemExit(f"Excel 文件不存在：{args.excel}")
    if args.workers < 1:
        raise SystemExit("--workers 必须大于 0。")

    # 准备输出目录
    args.output_dir.mkdir(parents=True, exist_ok=True)
    # 读取 Excel 中的图片 URL
    records = read_image_urls(args.excel, args.column)
    if not records:
        raise SystemExit(f"Excel 列 {args.column} 中没有可下载的 HTTP(S) 图片 URL。")

    # 多线程并发下载
    results: list[DownloadResult] = []
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = [executor.submit(download_one, record, args.output_dir, args.timeout) for record in records]
        for index, future in enumerate(as_completed(futures), start=1):
            result = future.result()
            results.append(result)
            print(f"[{index}/{len(futures)}] {result.status}: row {result.row_number} {result.path or result.error}")

    # 按行号排序结果
    results.sort(key=lambda item: (item.row_number, item.url))
    # 生成报告
    write_reports(results, args.metadata_dir)

    # 打印统计信息
    downloaded = sum(result.status == "downloaded" for result in results)
    skipped = sum(result.status == "skipped" for result in results)
    failed = sum(result.status == "failed" for result in results)
    print(f"完成：downloaded={downloaded}, skipped={skipped}, failed={failed}")
    print(f"原图目录：{args.output_dir}")
    print(f"下载报告：{args.metadata_dir / 'download_report.csv'}")


if __name__ == "__main__":
    main()
