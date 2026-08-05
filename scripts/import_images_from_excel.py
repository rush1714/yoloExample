from __future__ import annotations

import argparse
import csv
import hashlib
import json
import mimetypes
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from PIL import Image

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = Path("/Users/guobiao/Downloads/8e96894159cc584f0c7a27faaa4acc45.xlsx")
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "datasets" / "softcare" / "raw" / "images"
DEFAULT_METADATA_DIR = PROJECT_ROOT / "datasets" / "softcare" / "raw" / "metadata"
DEFAULT_COLUMN = "整改后图片URL"
IMAGE_EXTENSIONS = {".bmp", ".jpeg", ".jpg", ".png", ".tif", ".tiff", ".webp"}
CONTENT_TYPE_EXTENSIONS = {
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
    "image/bmp": ".bmp",
    "image/tiff": ".tiff",
}


@dataclass(frozen=True)
class ImageRecord:
    row_number: int
    url: str


@dataclass(frozen=True)
class DownloadResult:
    row_number: int
    url: str
    status: str
    path: str = ""
    error: str = ""


def find_column_index(headers: list[object], column_name: str) -> int:
    normalized = [str(value).strip() if value is not None else "" for value in headers]
    if column_name not in normalized:
        available = ", ".join(header for header in normalized if header)
        raise ValueError(f"Excel 中找不到列：{column_name}。当前列：{available}")
    return normalized.index(column_name)


def split_urls(value: object) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    for separator in ("\n", "\r", ";", "，", ","):
        text = text.replace(separator, " ")
    return [part.strip() for part in text.split() if part.strip()]


def read_image_urls(excel_path: Path, column_name: str) -> list[ImageRecord]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = list(next(rows))
    except StopIteration as exc:
        raise ValueError("Excel 是空文件。") from exc

    column_index = find_column_index(headers, column_name)
    records: list[ImageRecord] = []
    seen_urls: set[str] = set()
    for row_number, row in enumerate(rows, start=2):
        value = row[column_index] if column_index < len(row) else None
        for url in split_urls(value):
            parsed = urlparse(url)
            if parsed.scheme not in {"http", "https"} or not parsed.netloc:
                continue
            if url in seen_urls:
                continue
            seen_urls.add(url)
            records.append(ImageRecord(row_number=row_number, url=url))
    return records


def extension_from_response(url: str, content_type: str | None) -> str:
    parsed_suffix = Path(urlparse(url).path).suffix.lower()
    if parsed_suffix in IMAGE_EXTENSIONS:
        return parsed_suffix
    if content_type:
        media_type = content_type.split(";", 1)[0].strip().lower()
        if media_type in CONTENT_TYPE_EXTENSIONS:
            return CONTENT_TYPE_EXTENSIONS[media_type]
        guessed = mimetypes.guess_extension(media_type)
        if guessed in IMAGE_EXTENSIONS:
            return guessed
    return ".jpg"


def make_filename(record: ImageRecord, extension: str) -> str:
    digest = hashlib.sha1(record.url.encode("utf-8")).hexdigest()[:12]
    return f"row{record.row_number:05d}_{digest}{extension}"


def download_one(record: ImageRecord, output_dir: Path, timeout: int) -> DownloadResult:
    try:
        request = Request(record.url, headers={"User-Agent": "softcare-yolo-demo/0.1"})
        with urlopen(request, timeout=timeout) as response:
            if response.status != 200:
                return DownloadResult(record.row_number, record.url, "failed", error=f"HTTP {response.status}")
            content_type = response.headers.get("Content-Type")
            extension = extension_from_response(record.url, content_type)
            target = output_dir / make_filename(record, extension)
            if target.exists():
                return DownloadResult(record.row_number, record.url, "skipped", path=str(target))
            content = response.read()

        if not content:
            return DownloadResult(record.row_number, record.url, "failed", error="empty response")

        target.write_bytes(content)
        try:
            with Image.open(target) as image:
                image.verify()
        except Exception as exc:
            target.unlink(missing_ok=True)
            return DownloadResult(record.row_number, record.url, "failed", error=f"invalid image: {exc}")
        return DownloadResult(record.row_number, record.url, "downloaded", path=str(target))
    except HTTPError as exc:
        return DownloadResult(record.row_number, record.url, "failed", error=f"HTTP {exc.code}")
    except URLError as exc:
        return DownloadResult(record.row_number, record.url, "failed", error=str(exc.reason))
    except Exception as exc:
        return DownloadResult(record.row_number, record.url, "failed", error=str(exc))


def write_reports(results: list[DownloadResult], metadata_dir: Path) -> None:
    metadata_dir.mkdir(parents=True, exist_ok=True)
    json_path = metadata_dir / "download_report.json"
    csv_path = metadata_dir / "download_report.csv"

    rows = [result.__dict__ for result in results]
    json_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    with csv_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=["row_number", "url", "status", "path", "error"])
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="从 Excel 的图片 URL 列下载未标注原图。")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="Excel 文件路径")
    parser.add_argument("--column", default=DEFAULT_COLUMN, help="图片 URL 列名")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="原图保存目录")
    parser.add_argument("--metadata-dir", type=Path, default=DEFAULT_METADATA_DIR, help="下载报告目录")
    parser.add_argument("--workers", type=int, default=8, help="并发下载线程数")
    parser.add_argument("--timeout", type=int, default=30, help="单张图片下载超时时间（秒）")
    args = parser.parse_args()

    if not args.excel.is_file():
        raise SystemExit(f"Excel 文件不存在：{args.excel}")
    if args.workers < 1:
        raise SystemExit("--workers 必须大于 0。")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    records = read_image_urls(args.excel, args.column)
    if not records:
        raise SystemExit(f"Excel 列 {args.column} 中没有可下载的 HTTP(S) 图片 URL。")

    results: list[DownloadResult] = []
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = [executor.submit(download_one, record, args.output_dir, args.timeout) for record in records]
        for index, future in enumerate(as_completed(futures), start=1):
            result = future.result()
            results.append(result)
            print(f"[{index}/{len(futures)}] {result.status}: row {result.row_number} {result.path or result.error}")

    results.sort(key=lambda item: (item.row_number, item.url))
    write_reports(results, args.metadata_dir)

    downloaded = sum(result.status == "downloaded" for result in results)
    skipped = sum(result.status == "skipped" for result in results)
    failed = sum(result.status == "failed" for result in results)
    print(f"完成：downloaded={downloaded}, skipped={skipped}, failed={failed}")
    print(f"原图目录：{args.output_dir}")
    print(f"下载报告：{args.metadata_dir / 'download_report.csv'}")


if __name__ == "__main__":
    main()
